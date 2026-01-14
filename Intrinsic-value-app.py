# Intrinsic-value-app.py
# Streamlit app (Simple UX): Buffett-style intrinsic value (simplified DCF)
# - Excel template is OPTIONAL
# - If a bundled template exists at ./assets/intrinsic_value_template.xlsx, the app uses it automatically
# - Otherwise, user can upload a template (optional) to enable "download filled workbook"
#
# Persona-friendly defaults:
# - Terminal method defaults to Exit Multiple (simple)
# - Gordon Growth is hidden under Advanced
# - Only a handful of inputs shown in Simple mode

import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional

import streamlit as st

# Optional deps for template output
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # type: ignore


# -----------------------------
# Optional Yahoo fetch (yfinance)
# -----------------------------
def fetch_from_yahoo(ticker: str) -> Dict[str, float]:
    """
    Best-effort pull from Yahoo via yfinance:
      - sharesOutstanding
      - cash & equivalents (balance sheet)
      - total debt (balance sheet)
      - FCF0 = Operating CF - CapEx (cashflow)
    Coverage varies by ticker. Always review/override.
    """
    try:
        import yfinance as yf
    except Exception as e:
        raise RuntimeError("Missing dependency: yfinance. Install with `pip install yfinance`.") from e

    tk = yf.Ticker(ticker)

    out: Dict[str, float] = {}

    # shares from info
    try:
        info = tk.info or {}
        shares = info.get("sharesOutstanding")
        if shares is not None:
            out["shares"] = float(shares)
    except Exception:
        pass

    # balance sheet for cash/debt
    try:
        bs = tk.balance_sheet
        if bs is not None and not bs.empty:
            latest = bs.columns[0]
            idx = {str(i).strip(): i for i in bs.index}

            def pick(keys):
                for k in keys:
                    if k in idx:
                        return idx[k]
                return None

            cash_row = pick([
                "Cash And Cash Equivalents",
                "Cash And Cash Equivalents At Carrying Value",
                "Cash",
            ])
            debt_row = pick([
                "Total Debt",
                "Long Term Debt",
                "Short Long Term Debt",
            ])

            if cash_row is not None:
                out["cash"] = float(bs.loc[cash_row, latest])
            if debt_row is not None:
                out["debt"] = float(bs.loc[debt_row, latest])
    except Exception:
        pass

    # cashflow for FCF
    try:
        cf = tk.cashflow
        if cf is not None and not cf.empty:
            latest = cf.columns[0]
            idx = {str(i).strip(): i for i in cf.index}

            def pick(keys):
                for k in keys:
                    if k in idx:
                        return idx[k]
                return None

            ocf_row = pick([
                "Total Cash From Operating Activities",
                "Operating Cash Flow",
                "Net Cash Provided By Operating Activities",
            ])
            capex_row = pick([
                "Capital Expenditures",
                "Capital Expenditure",
            ])

            if ocf_row is not None and capex_row is not None:
                ocf = float(cf.loc[ocf_row, latest])
                capex = float(cf.loc[capex_row, latest])
                out["fcf0"] = ocf - capex
    except Exception:
        pass

    return out


# -----------------------------
# DCF Model
# -----------------------------
@dataclass
class Inputs:
    ticker: Optional[str] = None
    currency: str = "USD"

    fcf0: Optional[float] = None
    shares: Optional[float] = None
    cash: float = 0.0
    debt: float = 0.0

    years: int = 10
    growth: float = 0.06          # single growth for Y1-10
    discount_rate: float = 0.10   # required return

    terminal_method: str = "ExitMultiple"  # ExitMultiple (simple) / Gordon (advanced)
    exit_multiple: float = 15.0
    terminal_growth: float = 0.03

    mos: float = 0.30


def calc_dcf(inp: Inputs) -> Dict[str, float]:
    if inp.fcf0 is None:
        raise ValueError("FCF is required. Enter it manually or fetch via ticker.")
    if inp.shares is None or inp.shares <= 0:
        raise ValueError("Shares outstanding is required and must be > 0.")
    if inp.discount_rate <= 0:
        raise ValueError("Required return must be > 0.")
    if inp.terminal_method == "Gordon" and inp.discount_rate <= inp.terminal_growth:
        raise ValueError("For Gordon method, required return must be greater than terminal growth.")

    years = max(5, min(int(inp.years), 10))

    pv_cashflows = 0.0
    fcf = inp.fcf0
    for t in range(1, years + 1):
        fcf *= (1 + inp.growth)
        pv_cashflows += fcf / ((1 + inp.discount_rate) ** t)

    # terminal value at end of year N
    if inp.terminal_method == "Gordon":
        tv = fcf * (1 + inp.terminal_growth) / (inp.discount_rate - inp.terminal_growth)
    else:
        tv = fcf * inp.exit_multiple

    pv_tv = tv / ((1 + inp.discount_rate) ** years)

    ev = pv_cashflows + pv_tv
    net_debt = inp.debt - inp.cash
    equity_value = ev - net_debt
    ivps = equity_value / inp.shares
    buy_price = ivps * (1 - inp.mos)

    return {
        "enterprise_value": ev,
        "equity_value": equity_value,
        "intrinsic_value_per_share": ivps,
        "buy_price_mos": buy_price,
    }


# -----------------------------
# Optional Excel output
# -----------------------------
def load_bundled_template() -> Optional[bytes]:
    p = Path("assets") / "intrinsic_value_template.xlsx"
    return p.read_bytes() if p.exists() else None


def write_filled_workbook(template_bytes: bytes, inp: Inputs, res: Dict[str, float]) -> bytes:
    if load_workbook is None:
        raise RuntimeError("openpyxl not installed; cannot create Excel output.")

    with tempfile.TemporaryDirectory() as d:
        dpath = Path(d)
        tpath = dpath / "template.xlsx"
        tpath.write_bytes(template_bytes)

        wb = load_workbook(tpath)
        if "Inputs" not in wb.sheetnames or "Model" not in wb.sheetnames:
            raise ValueError("Template missing required sheets: Inputs and Model.")

        wsI = wb["Inputs"]
        wsM = wb["Model"]

        # These cell positions match the template we generated earlier.
        wsI["B5"].value = inp.ticker or ""
        wsI["B7"].value = inp.cash
        wsI["B8"].value = inp.debt
        wsI["B9"].value = inp.shares
        wsI["B13"].value = inp.fcf0
        wsI["B14"].value = inp.years
        wsI["B15"].value = inp.discount_rate
        wsI["B16"].value = "Single"
        wsI["B17"].value = inp.growth
        wsI["B34"].value = inp.terminal_method
        wsI["B35"].value = inp.terminal_growth
        wsI["B36"].value = inp.exit_multiple
        wsI["B39"].value = inp.mos

        # Write key outputs
        wsM["B29"].value = res["enterprise_value"]
        wsM["B31"].value = res["equity_value"]
        wsM["B32"].value = res["intrinsic_value_per_share"]
        wsM["B33"].value = res["buy_price_mos"]

        outp = dpath / "filled_intrinsic_value.xlsx"
        wb.save(outp)
        return outp.read_bytes()


# -----------------------------
# Streamlit UI
# -----------------------------
st.set_page_config(page_title="Intrinsic Value (Simple)", layout="wide")
st.title("Intrinsic Value Calculator (Simple)")

st.markdown(
    "Enter a few simple inputs to estimate intrinsic value per share.\n\n"
    "- **Excel is NOT required** unless you want a filled workbook download.\n"
    "- If you have a template bundled at `assets/intrinsic_value_template.xlsx`, upload is not needed."
)

bundled = load_bundled_template()
template_upload = None

with st.expander("Excel download (optional)"):
    if bundled is not None:
        st.success("Bundled template found: assets/intrinsic_value_template.xlsx")
    else:
        st.info("No bundled template found. Upload a template to enable Excel download.")
        template_upload = st.file_uploader("Upload intrinsic_value_template.xlsx (optional)", type=["xlsx"])

st.divider()

c1, c2, c3, c4 = st.columns(4)
with c1:
    ticker = st.text_input("Ticker (optional)", value="")
    fetch = st.checkbox("Auto-fill from Yahoo (optional)", value=False)

with c2:
    fcf0 = st.number_input("FCF (last year)", value=0.0, format="%.2f")
    shares = st.number_input("Shares outstanding", value=0.0, format="%.2f")

with c3:
    cash = st.number_input("Cash", value=0.0, format="%.2f")
    debt = st.number_input("Debt", value=0.0, format="%.2f")

with c4:
    growth = st.number_input("Growth % (next 10 yrs)", value=0.06, min_value=-0.50, max_value=1.00, step=0.01, format="%.4f")
    discount = st.number_input("Required return %", value=0.10, min_value=0.0, max_value=1.0, step=0.005, format="%.4f")

c5, c6, c7 = st.columns(3)
with c5:
    exit_multiple = st.number_input("Exit multiple", value=15.0, min_value=0.0, max_value=100.0, step=0.5, format="%.2f")
with c6:
    mos = st.number_input("Margin of safety", value=0.30, min_value=0.0, max_value=0.95, step=0.05, format="%.2f")
with c7:
    years = st.slider("Forecast years", min_value=5, max_value=10, value=10)

with st.expander("Advanced (optional)"):
    st.caption("If unsure, ignore this section. Simple mode uses Exit Multiple.")
    terminal_method = st.selectbox("Terminal method", ["ExitMultiple", "Gordon"], index=0)
    terminal_growth = st.number_input("Terminal growth (Gordon only)", value=0.03, min_value=-0.10, max_value=0.20, step=0.005, format="%.4f")

run = st.button("Calculate", type="primary")

if run:
    inp = Inputs(
        ticker=ticker.strip() or None,
        fcf0=None if fcf0 == 0 else float(fcf0),
        shares=None if shares == 0 else float(shares),
        cash=float(cash),
        debt=float(debt),
        years=int(years),
        growth=float(growth),
        discount_rate=float(discount),
        terminal_method=terminal_method,
        exit_multiple=float(exit_multiple),
        terminal_growth=float(terminal_growth),
        mos=float(mos),
    )

    if fetch:
        if not inp.ticker:
            st.error("Enter a ticker to auto-fill from Yahoo.")
            st.stop()
        try:
            y = fetch_from_yahoo(inp.ticker)
        except Exception as e:
            st.error(f"Yahoo fetch failed: {e}")
            st.stop()

        # fill missing only
        if inp.fcf0 is None and "fcf0" in y:
            inp.fcf0 = y["fcf0"]
        if inp.shares is None and "shares" in y:
            inp.shares = y["shares"]
        if inp.cash == 0.0 and "cash" in y:
            inp.cash = y["cash"]
        if inp.debt == 0.0 and "debt" in y:
            inp.debt = y["debt"]

    try:
        res = calc_dcf(inp)
    except Exception as e:
        st.error(str(e))
        st.stop()

    st.success("Valuation completed.")
    o1, o2, o3, o4 = st.columns(4)
    o1.metric("Enterprise Value", f"{res['enterprise_value']:,.2f}")
    o2.metric("Equity Value", f"{res['equity_value']:,.2f}")
    o3.metric("Intrinsic Value / Share", f"{res['intrinsic_value_per_share']:,.4f}")
    o4.metric("Buy Price (MOS)", f"{res['buy_price_mos']:,.4f}")

    # Excel output (optional)
    template_bytes = bundled or (template_upload.getvalue() if template_upload is not None else None)
    if template_bytes is not None:
        try:
            filled = write_filled_workbook(template_bytes, inp, res)
            st.download_button(
                "Download filled workbook",
                data=filled,
                file_name="filled_intrinsic_value.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.warning(f"Could not create Excel output: {e}")
    else:
        st.info("Excel download is disabled (no template provided).")
